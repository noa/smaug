"""Operational CLI commands: travel, expense, invoice, note management."""

from decimal import Decimal
from pathlib import Path

from ..store import ProjectStore
from ._util import Anonymizer, Colors, color, parse_date_input, resolve_personnel_name


def cmd_travel(store: ProjectStore, args) -> None:
    """Manage travel budget items."""

    config_path = Path(args.data_dir) / "projects" / "travel_config.yaml"

    if args.action == "list":
        # Reload to ensure fresh data
        store.load_travel_config()

        # Filter by project if specified
        if args.project:
            items = store.get_project_travel(args.project)
            title = f"Travel: {args.project}"
        else:
            items = store._travel
            title = "All Travel Items"

        print(f"\n=== {title} ===\n")
        print(
            f"{'Date':<12} {'Project':<10} {'Traveler':<20} {'Description':<30} {'Amount':>10} {'Status':<12}"
        )
        print("-" * 99)

        for item in items:
            travel_date_str = item.date.strftime("%Y-%m-%d") if item.date else "-"
            print(
                f"{travel_date_str:<12} {item.project_id:<10} {Anonymizer.anonymize(item.traveler or '-'):<20} {item.description:<30} ${item.amount:>9,.2f} {item.status.value:<12}"
            )

        print("-" * 99)

    elif args.action == "add":
        # Parse date
        from datetime import date, datetime

        travel_date = None
        if args.date:
            try:
                travel_date = datetime.strptime(args.date.strip(), "%Y-%m-%d").date()
            except ValueError:
                parsed_ym = parse_date_input(args.date)
                parts = parsed_ym.split("-")
                travel_date = date(int(parts[0]), int(parts[1]), 1)
        else:
            travel_date = date.today()

        # Resolve traveler (allow missing for external travelers)
        traveler_name = args.traveler
        if args.traveler:
            tracker = store.get_personnel_tracker()
            personnel = tracker.get_all_personnel()
            resolved_name, error = resolve_personnel_name(
                args.traveler, personnel, allow_missing=True
            )

            if error:
                print(f"Error: {error}")
                return

            traveler_name = resolved_name
            if resolved_name != args.traveler:
                print(
                    f"Resolved traveler '{args.traveler}' to: {Anonymizer.anonymize(resolved_name)}"
                )

        new_item = {
            "project": args.project,
            "description": args.description,
            "date": travel_date,
            "amount": float(args.amount),
            "traveler": traveler_name,
            "status": "estimated",
        }

        from ..yaml_utils import git_commit_change, yaml_transaction

        with yaml_transaction(config_path) as existing_data:
            existing_data.setdefault("travel", []).append(new_item)

        print(f"Added travel item to {args.project}: {args.description} (${args.amount})")
        git_commit_change(
            args.data_dir, f"travel-add: {args.project} - {args.description} (${args.amount})"
        )

    elif args.action == "actualize":
        if not config_path.exists():
            print("No travel configuration found.")
            return

        from ..yaml_utils import git_commit_change, yaml_transaction

        try:
            with yaml_transaction(config_path) as data:
                travel_list = data.get("travel", [])

                # Find item matching description (partial match) and project
                matches = []
                for idx, entry in enumerate(travel_list):
                    if (
                        entry.get("project") == args.project
                        and args.description.lower() in entry.get("description", "").lower()
                    ):
                        matches.append((idx, entry))

                if not matches:
                    raise ValueError(
                        f"No travel item found for {args.project} matching '{args.description}'"
                    )

                if len(matches) > 1:
                    match_list = "\n".join(
                        f"  - {entry['date']} {entry['description']} (${entry['amount']})"
                        for _, entry in matches
                    )
                    raise ValueError(
                        f"Multiple matches found for '{args.description}':\n{match_list}\nPlease be more specific."
                    )

                idx, entry = matches[0]

                old_amount = entry.get("amount", 0)
                entry["status"] = "actualized"
                if args.amount:
                    entry["amount"] = float(args.amount)
        except ValueError as e:
            print(f"Error: {e}")
            return

        amount_str = f"${entry['amount']}" if args.amount else f"${old_amount} (unchanged)"
        print(f"Actualized: {entry['description']} — {amount_str}")
        git_commit_change(
            args.data_dir, f"travel-actualize: {args.project} - {entry['description']}"
        )


def cmd_expense(store: ProjectStore, args) -> None:
    """Manage recurrent and one-time expenses."""

    config_path = Path(args.data_dir) / "projects" / "purchases_config.yaml"

    if args.action == "list":
        store.load_purchases_config()

        if args.project:
            items = store.get_project_expenses(args.project)
            title = f"Expenses: {args.project}"
        else:
            items = store._expenses
            title = "All Expenses"

        print(f"\n=== {title} ===")
        print(
            f"{'Date/Start':<12} {'End':<12} {'Project':<10} {'Description':<30} {'Category':<15} {'Amount':>10} {'Freq':<10}"
        )
        print("-" * 105)

        for item in items:
            date_str = (
                item.date.strftime("%Y-%m-%d")
                if item.date
                else (item.start_date.strftime("%Y-%m-%d") if item.start_date else "-")
            )
            end_str = item.end_date.strftime("%Y-%m-%d") if item.end_date else "-"
            freq = "Monthly" if item.is_recurring else "One-time"

            print(
                f"{date_str:<12} {end_str:<12} {item.project_id:<10} {item.description:<30} {item.category:<15} ${item.amount:>9,.2f} {freq:<10}"
            )

        print("-" * 105)

    elif args.action == "add":
        new_item = {
            "project": args.project,
            "description": args.description,
            "amount": float(args.amount),
            "category": args.category,
        }

        if args.date:
            new_item["date"] = parse_date_input(args.date)
            print(
                f"Added one-time expense to {args.project}: {args.description} (${args.amount}) on {new_item['date']}"
            )
        elif args.start and args.end:
            new_item["start"] = parse_date_input(args.start)
            new_item["end"] = parse_date_input(args.end)
            print(
                f"Added recurring expense to {args.project}: {args.description} (${args.amount}/mo) from {new_item['start']} to {new_item['end']}"
            )
        else:
            # Default to recurring from today through project end
            from datetime import date

            project_data = store.get_project(args.project)
            if not project_data or not project_data.project.end_date:
                print(
                    f"Error: Could not determine end date for project {args.project}. Please specify --end or --date."
                )
                return

            new_item["start"] = date.today().replace(day=1)
            new_item["end"] = project_data.project.end_date
            print(
                f"Added recurring expense to {args.project}: {args.description} (${args.amount}/mo) from {new_item['start']} to {new_item['end']}"
            )

        from ..yaml_utils import git_commit_change, yaml_transaction

        with yaml_transaction(config_path) as existing_data:
            existing_data.setdefault("items", []).append(new_item)

        git_commit_change(
            args.data_dir, f"expense-add: {args.project} - {args.description} (${args.amount})"
        )


def cmd_invoice(store: ProjectStore, args) -> None:
    """Manage and validate invoices."""

    action = args.action if hasattr(args, "action") else "list"
    project_filter = getattr(args, "project", None)

    if action == "list":
        # List all invoices
        projects = [project_filter] if project_filter else store.list_projects()

        print(f"\n{color('=== Invoices ===', Colors.BOLD + Colors.CYAN)}\n")
        print(
            f"{'Project':<10} {'Invoice #':<14} {'Period':<24} {'Current':>14} {'Cumulative':>16}"
        )
        print(color("-" * 82, Colors.DIM))

        for project_id in projects:
            invoices = store.get_project_invoices(project_id)
            if not invoices:
                continue

            for inv in invoices:
                period_str = (
                    f"{inv.period_start.strftime('%b %d')} - {inv.period_end.strftime('%b %d, %Y')}"
                )
                print(
                    f"{color(project_id, Colors.BOLD):<19} {inv.invoice_number:<14} {period_str:<24} ${inv.current_expense:>12,.2f} ${inv.cumulative_expense:>14,.2f}"
                )

        if not any(store.get_project_invoices(p) for p in projects):
            print("No invoices found.")

    elif action == "validate":
        # Validate invoices against spending reports
        projects = [project_filter] if project_filter else store.list_projects()

        print(f"\n{color('=== Invoice Validation ===', Colors.BOLD + Colors.CYAN)}\n")

        threshold = Decimal("0.01")  # Allow 1 cent tolerance for rounding

        for project_id in projects:
            invoices = store.get_project_invoices(project_id)
            data = store.get_project(project_id)

            if not invoices:
                continue

            print(f"{color(project_id, Colors.BOLD + Colors.YELLOW)}")

            if not data or not data.spending:
                print(f"  {color('⚠ No spending reports found for comparison', Colors.YELLOW)}")
                continue

            for inv in invoices:
                # Find the spending report for the invoice period
                # Invoice period_end is the last day of the month, find report for that month
                inv_year = inv.period_end.year
                inv_month = inv.period_end.month

                # Calculate cumulative spending from all reports up to and including this month
                reports_up_to = [
                    r for r in data.spending if (r.year, r.month) <= (inv_year, inv_month)
                ]

                if not reports_up_to:
                    print(
                        f"  Invoice {inv.invoice_number} ({inv.period_end.strftime('%b %Y')}): {color('No matching reports', Colors.YELLOW)}"
                    )
                    continue

                # Get the cumulative from the latest report up to this period
                latest_report = max(reports_up_to, key=lambda r: (r.year, r.month))
                report_cumulative = latest_report.total_spent

                # Compare
                diff = abs(inv.cumulative_expense - report_cumulative)
                pct_diff = (
                    (diff / inv.cumulative_expense * 100) if inv.cumulative_expense > 0 else 0
                )

                period_str = inv.period_end.strftime("%b %Y")

                if diff <= threshold:
                    status = color("✓", Colors.GREEN)
                    detail = f"${inv.cumulative_expense:,.2f}"
                elif pct_diff < 5:
                    status = color("~", Colors.YELLOW)
                    detail = f"Invoice: ${inv.cumulative_expense:,.2f}, Reports: ${report_cumulative:,.2f} (diff: ${diff:,.2f}, {pct_diff:.1f}%)"
                else:
                    status = color("✗", Colors.RED)
                    detail = f"Invoice: ${inv.cumulative_expense:,.2f}, Reports: ${report_cumulative:,.2f} (diff: ${diff:,.2f}, {pct_diff:.1f}%)"

                print(f"  {status} Invoice {inv.invoice_number} ({period_str}): {detail}")

                # Also check category breakdown if verbose
                if hasattr(args, "verbose") and args.verbose:
                    print("    Categories in invoice:")
                    for cat, amount in sorted(inv.categories.items()):
                        print(f"      {cat}: ${amount:,.2f}")

            print()  # Blank line between projects

    elif action == "import":
        from ._import import cmd_invoice_import

        cmd_invoice_import(store, args)


def cmd_note(store: ProjectStore, args) -> None:
    """Manage per-project notes."""
    from ..notes import add_note, import_note, list_notes, remove_note, show_note

    project_id = args.project

    if args.action == "list":
        notes = list_notes(args.data_dir, project_id)
        if not notes:
            print(f"No notes for {project_id}.")
            return

        print(f"\n=== Notes: {project_id} ===")
        print(f"{'#':<4} {'Date':<12} {'Title':<50} {'Tags'}")
        print("-" * 85)
        for i, note in enumerate(notes, 1):
            date_str = note.created.strftime("%Y-%m-%d")
            tags_str = ", ".join(note.tags) if note.tags else ""
            print(f"{i:<4} {date_str:<12} {note.title:<50} {tags_str}")

    elif args.action == "show":
        content, error = show_note(args.data_dir, project_id, args.identifier)
        if error:
            print(f"Error: {error}")
            return
        print(content)

    elif args.action == "add":
        import os
        import subprocess
        import tempfile

        title = args.title
        content = getattr(args, "message", None) or ""
        tags = [t.strip() for t in args.tags.split(",")] if getattr(args, "tags", None) else []

        if not content:
            # Open $EDITOR
            editor = os.environ.get("EDITOR", "vi")
            with tempfile.NamedTemporaryFile(suffix=".md", mode="w+", delete=False) as tmp:
                tmp.write(f"# {title}\n\n")
                tmp.flush()
                tmp_path = tmp.name
            try:
                subprocess.call([editor, tmp_path])
                with open(tmp_path) as f:
                    content = f.read()
            finally:
                Path(tmp_path).unlink(missing_ok=True)

        if not content.strip():
            print("Empty note, not saved.")
            return

        filepath = add_note(args.data_dir, project_id, title, content, tags=tags)
        print(f"Note saved: {filepath.name}")

    elif args.action in ("remove", "rm"):
        removed_title, error = remove_note(args.data_dir, project_id, args.identifier)
        if error:
            print(f"Error: {error}")
            return
        print(f"Removed note: {removed_title}")

    elif args.action == "import":
        source = Path(args.file)
        if not source.exists():
            print(f"File not found: {args.file}")
            return
        title_override = getattr(args, "title", None)
        tags = [t.strip() for t in args.tags.split(",")] if getattr(args, "tags", None) else []
        filepath = import_note(args.data_dir, project_id, source, title=title_override, tags=tags)
        print(f"Imported note: {filepath.name}")


def cmd_init(store: ProjectStore, args) -> None:
    """Scaffold ~/.smaug/ or custom data directory with standard files."""
    import shutil
    import subprocess

    data_dir = Path(args.data_dir)
    print(f"\nInitializing Smaug data directory at: {data_dir}\n")

    # Create directory structure
    projects_dir = data_dir / "projects"
    reports_sponsored_dir = data_dir / "reports" / "sponsored"
    reports_nonsponsored_dir = data_dir / "reports" / "non-sponsored"
    invoices_dir = data_dir / "reports" / "invoices"

    for d in (projects_dir, reports_sponsored_dir, reports_nonsponsored_dir, invoices_dir):
        d.mkdir(parents=True, exist_ok=True)
        print(f"  Created directory: {d.relative_to(data_dir.parent)}")

    # Copy templates from examples
    examples_dir = Path(__file__).parent.parent.parent.parent / "examples"
    templates = [
        ("rates.yaml", data_dir / "rates.yaml"),
        ("projects/manifest.yaml", projects_dir / "manifest.yaml"),
        ("projects/personnel_config.yaml", projects_dir / "personnel_config.yaml"),
        ("projects/travel_config.yaml", projects_dir / "travel_config.yaml"),
        ("projects/purchases_config.yaml", projects_dir / "purchases_config.yaml"),
    ]

    for src_rel, dest in templates:
        src = examples_dir / src_rel
        if src.exists() and not dest.exists():
            shutil.copy2(src, dest)
            print(f"  Copied template: {dest.name}")
        elif dest.exists():
            print(f"  Existing file kept: {dest.name}")

    # Set up Git change tracking
    git_dir = data_dir / ".git"
    if not git_dir.exists():
        try:
            subprocess.run(["git", "init"], cwd=str(data_dir), capture_output=True, check=True)
            subprocess.run(["git", "add", "."], cwd=str(data_dir), capture_output=True, check=True)
            subprocess.run(
                ["git", "commit", "-m", "Initial Smaug repository scaffolding"],
                cwd=str(data_dir),
                capture_output=True,
                check=True,
            )
            print("  Initialized Git repository and committed initial templates.")
        except Exception as e:
            print(f"  Warning: Failed to set up Git repository: {e}")
    else:
        print("  Git repository already initialized.")

    print("\nSmaug initialization complete!")
    print("Run 'smaug list' to verify that the template projects are loaded successfully.")


def cmd_health(store: ProjectStore, args) -> None:
    """Run comprehensive data integrity checks and display a dashboard."""
    print(f"\n{color('=== Smaug Data Integrity Dashboard ===', Colors.BOLD + Colors.CYAN)}\n")

    warnings = []

    # 1. Check missing spending report coverage (gaps)
    from datetime import datetime

    now = datetime.now()
    projects = store.list_projects()
    for project_id in projects:
        data = store.get_project(project_id)
        if not data or not data.spending:
            warnings.append(f"Project '{project_id}': No spending reports found.")
            continue
        periods = {(r.year, r.month) for r in data.spending}
        min_period = min(periods)
        max_period = max(periods)
        end_year, end_month = now.year, now.month
        if max_period > (end_year, end_month):
            end_year, end_month = max_period

        year, month = min_period
        missing_months = []
        while (year, month) <= (end_year, end_month):
            if (year, month) not in periods:
                missing_months.append(f"{datetime(year, month, 1).strftime('%B %Y')}")
            month += 1
            if month > 12:
                month = 1
                year += 1
        if missing_months:
            warnings.append(
                f"Project '{project_id}' has missing report gaps: {', '.join(missing_months)}."
            )

    # 2. Check missing invoices gaps (gaps in project invoices)
    for project_id in projects:
        invoices = store.get_project_invoices(project_id)
        data = store.get_project(project_id)
        # Check sponsored projects with budgets for invoices
        if data and data.project.project_type.value == "sponsored":
            if not invoices:
                warnings.append(f"Project '{project_id}': No invoices found.")
            else:
                # Check for month coverage from earliest invoice to now
                periods = {(inv.period_end.year, inv.period_end.month) for inv in invoices}
                min_period = min(periods)
                end_year, end_month = now.year, now.month
                year, month = min_period
                missing_invs = []
                while (year, month) <= (end_year, end_month):
                    if (year, month) not in periods:
                        missing_invs.append(f"{datetime(year, month, 1).strftime('%B %Y')}")
                    month += 1
                    if month > 12:
                        month = 1
                        year += 1
                if missing_invs:
                    warnings.append(
                        f"Project '{project_id}' has missing invoice gaps: {', '.join(missing_invs)}."
                    )

    # 3. Check ParseWarnings (unresolved invoices etc.)
    for pw in store.get_parse_warnings():
        warnings.append(f"Parse warning in {pw.file}: {pw.message}")

    # 4. Check Stale Salaries and overcommitment from personnel config
    config_path = Path(args.data_dir) / "projects" / "personnel_config.yaml"
    if config_path.exists():
        from ..projections import load_personnel_config

        try:
            _rates, config_personnel = load_personnel_config(config_path)

            # Check over-commitment: total effort across all projects > 100%
            for p in config_personnel:
                total_effort = sum(a.effort for a in p.assignments)
                if total_effort > 1.0:
                    warnings.append(
                        f"Person '{Anonymizer.anonymize(p.name)}' is overcommitted at {total_effort * 100:.0f}% effort across projects."
                    )

                # Check stale salaries: warn if salary is 0 or potentially stale
                if p.annual_salary <= 0:
                    warnings.append(
                        f"Person '{Anonymizer.anonymize(p.name)}' has a zero or missing annual salary in personnel_config.yaml."
                    )

                # Check orphan assignments: projects in assignments not in manifest
                for a in p.assignments:
                    if a.project not in projects:
                        warnings.append(
                            f"Person '{Anonymizer.anonymize(p.name)}' is assigned to orphan project '{a.project}' which does not exist in manifest.yaml."
                        )
        except Exception as e:
            warnings.append(f"Failed to check personnel config integrity: {e}")

    # 5. Check budget alert thresholds (Feature E)
    manifest_path = Path(args.data_dir) / "projects" / "manifest.yaml"
    if manifest_path.exists():
        try:
            with open(manifest_path, encoding="utf-8") as f:
                import yaml

                manifest_data = yaml.safe_load(f) or {}

            for project_id in projects:
                proj_conf = (
                    manifest_data.get("projects", {}).get(project_id)
                    or manifest_data.get("discretionary", {}).get(project_id)
                    or {}
                )
                alerts_conf = proj_conf.get("alerts", {})
                if alerts_conf:
                    # Load status and stopwork forecast
                    from ..api import SmaugAPI

                    api = SmaugAPI(args.data_dir)
                    status_res = api.project_status(project_id)
                    stopwork_res = api.stopwork_forecast(project_id)

                    if "error" not in status_res:
                        # Budget remaining pct check
                        budget_remaining_pct = alerts_conf.get("budget_remaining_pct")
                        if budget_remaining_pct and "pct_remaining" in status_res:
                            pct_rem = status_res["pct_remaining"]
                            if pct_rem < budget_remaining_pct:
                                warnings.append(
                                    f"ALERT: Project '{project_id}' budget remaining is {pct_rem:.1f}%, which is below threshold of {budget_remaining_pct}%."
                                )

                        # Stopwork date forecast check
                        months_to_stopwork = alerts_conf.get("months_to_stopwork")
                        if months_to_stopwork and "months_remaining" in stopwork_res:
                            months_rem = stopwork_res["months_remaining"]
                            if months_rem < months_to_stopwork:
                                warnings.append(
                                    f"ALERT: Project '{project_id}' stop-work date forecast is {months_rem:.1f} months away, which is below threshold of {months_to_stopwork} months."
                                )
        except Exception as e:
            warnings.append(f"Failed to parse alert thresholds: {e}")

    # Display results
    if not warnings:
        print(
            f"  {color('✓ All checks passed perfectly! Your data is completely healthy and synchronized.', Colors.GREEN)}"
        )
    else:
        print(
            f"  {color(f'Found {len(warnings)} data integrity issue(s):', Colors.YELLOW + Colors.BOLD)}\n"
        )
        for idx, w in enumerate(warnings, 1):
            if "ALERT" in w:
                print(f"  {idx:>2}. {color('⚠', Colors.RED)} {color(w, Colors.RED)}")
            else:
                print(f"  {idx:>2}. {color('⚠', Colors.YELLOW)} {w}")
    print()


def cmd_history(store: ProjectStore, args) -> None:
    """Show git history of changes in the data directory."""
    import subprocess

    data_dir = args.data_dir
    git_dir = Path(data_dir) / ".git"
    if not git_dir.exists():
        print("Git change-tracking is not initialized for this data directory.")
        print("Run 'smaug init' to set up git change tracking.")
        return

    try:
        # Wrap git log nicely
        result = subprocess.run(
            ["git", "log", "-n", "20", "--oneline", "--decorate", "--color"],
            cwd=str(data_dir),
            capture_output=True,
            text=True,
            check=True,
        )
        print("\n=== Smaug Configuration History ===\n")
        print(result.stdout)
    except Exception as e:
        print(f"Error reading history: {e}")


def cmd_undo(store: ProjectStore, args) -> None:
    """Revert the last modification in the data directory."""
    import subprocess

    data_dir = args.data_dir
    git_dir = Path(data_dir) / ".git"
    if not git_dir.exists():
        print("Git change-tracking is not initialized for this data directory.")
        return

    try:
        # Check if there are uncommitted changes
        status = subprocess.run(
            ["git", "status", "--porcelain"],
            cwd=str(data_dir),
            capture_output=True,
            text=True,
            check=True,
        )
        if status.stdout.strip():
            print("Warning: You have uncommitted changes. Stashing them first...")
            subprocess.run(["git", "stash"], cwd=str(data_dir), capture_output=True)

        # Revert the last commit
        result = subprocess.run(
            ["git", "revert", "HEAD", "--no-edit"],
            cwd=str(data_dir),
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            print("Successfully reverted the last command modification!")
            # Show the new HEAD commit
            log = subprocess.run(
                ["git", "log", "-1", "--oneline"], cwd=str(data_dir), capture_output=True, text=True
            )
            print(f"Current state: {log.stdout.strip()}")
        else:
            print("Revert failed. There might not be any commits to revert, or conflict occurred.")
            print(result.stderr)
    except Exception as e:
        print(f"Error executing undo: {e}")


def cmd_export(store: ProjectStore, args) -> None:
    """Export the project spend plan to an Excel file using openpyxl."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from ..api import SmaugAPI

    project_id = args.project
    filename = args.filename

    api = SmaugAPI(args.data_dir)
    res = api.spending_projection(project_id, months=36)

    if "error" in res:
        print(f"Error generating projection: {res['error']}")
        return

    projections = res["projections"]

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"Spend Plan {project_id}"

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Smaug Spend Plan - Project: {project_id}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ws.row_dimensions[1].height = 40

    # Headers
    headers = ["Month", "Salary", "Fringe", "Travel", "Compute", "Equip", "Other", "IDC", "Total"]
    ws.append(headers)
    ws.row_dimensions[2].height = 24

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 10):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Append data
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_right = Alignment(horizontal="right")

    for row_idx, p in enumerate(projections, 3):
        row_data = [
            p["month"],
            p["salary"],
            p["fringe"],
            p["travel"],
            p["compute"],
            p["equipment"],
            p["other"],
            p["indirect"],
            p["total"],
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20

        # Style row
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        for col_idx in range(2, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "$#,##0.00"
            cell.alignment = align_right

        # Alternating row colors
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            for col_idx in range(1, 10):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Add border
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_num = col[0].column
        if col_num is None:
            continue
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filename)
    print(f"Successfully exported styled spend plan to: {filename}")
